"""Parse uploaded CSV/XLSX bytes into a normalized sheet structure."""

import csv
import io
import re
import sys
import zipfile
from dataclasses import dataclass, field

from openpyxl import load_workbook

from app.engine.headings import has_header

# A single cell can legitimately hold a pasted document; the stdlib default
# of 128 KB rejects those outright.
csv.field_size_limit(min(sys.maxsize, 2**31 - 1))


@dataclass
class Sheet:
    name: str
    headers: list[str]
    rows: list[list[str]] = field(default_factory=list)
    # Which character the file was split on, and the runner-up if the
    # split was a genuine toss-up, so a wrong guess can be surfaced
    delimiter: str = ""
    rival_delimiter: str | None = None


def read_file(filename: str, data: bytes) -> list[Sheet]:
    """Parse an upload into sheets. Every failure raises ValueError with a
    message written for someone who has never heard of an encoding."""
    lower = filename.lower()
    if not data.strip():
        raise ValueError("This file is empty — there's nothing to work with.")

    if lower.endswith(".xls"):
        raise ValueError(
            "Excel 97-2003 files (.xls) aren't supported. Open the file in Excel "
            "and use File → Save As → Excel Workbook (.xlsx), then try again."
        )
    if lower.endswith(".csv") or lower.endswith(".txt") or lower.endswith(".tsv"):
        sheets = [_read_csv(data)]
    elif lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        sheets = _read_xlsx(data)
    elif lower.endswith(".zip"):
        raise ValueError(
            "This is a ZIP folder, not a spreadsheet. Unzip it first (double-click "
            "it), then choose the .csv or .xlsx file inside."
        )
    else:
        raise ValueError(
            "This tool reads spreadsheet files: CSV (.csv) or Excel (.xlsx). "
            f"'{filename}' is something else — try exporting your data as CSV first."
        )

    if not any(s.rows for s in sheets):
        raise ValueError(
            "This file has column headings but no data rows underneath them, "
            "so there's nothing to process."
        )
    return sheets


def _cell(value) -> str:
    return "" if value is None else str(value)


def _normalize(headers: list[str], rows: list[list[str]]) -> Sheet:
    """Make every row exactly as wide as the header row. Short rows are
    padded; when a data row is wider than the headers, the headers are
    extended instead of hiding the extra cells (a hidden cell could carry
    PII the user never sees or redacts)."""
    width = max([len(headers), *(len(r) for r in rows)] if rows else [len(headers)])
    headers = headers + [f"extra_column_{i + 1}" for i in range(len(headers), width)]
    rows = [row + [""] * (width - len(row)) for row in rows]

    # Column selection is keyed by header name, so names must be unique and
    # non-empty or some columns could never be addressed for redaction.
    seen: dict[str, int] = {}
    unique = []
    for i, name in enumerate(headers):
        name = name.strip() or f"column_{i + 1}"
        count = seen.get(name, 0) + 1
        seen[name] = count
        unique.append(name if count == 1 else f"{name} ({count})")
    return unique, rows


def _decode(data: bytes) -> str:
    """Decode spreadsheet bytes, then normalize line endings.

    Excel writes UTF-16 for "Unicode Text", older exports are cp1252, and
    classic-Mac tools still emit CR-only line endings. latin-1 is the
    terminal fallback: it accepts any byte sequence, so decoding never
    fails outright.
    """
    text = None
    if data[:2] in (b"\xff\xfe", b"\xfe\xff"):
        try:
            text = data.decode("utf-16")
        except UnicodeDecodeError:
            text = None
    if text is None:
        # UTF-16 without a BOM shows up as every other byte being NUL.
        head = data[:4096]
        if head.count(0) > len(head) // 4:
            for encoding in ("utf-16-le", "utf-16-be"):
                try:
                    candidate = data.decode(encoding)
                except UnicodeDecodeError:
                    continue
                if "\x00" not in candidate:
                    text = candidate
                    break
    if text is None:
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = data.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
    if text is None:  # pragma: no cover - latin-1 accepts anything
        raise AssertionError("unreachable")
    return text.replace("\r\n", "\n").replace("\r", "\n")


def _sniff_delimiter(text: str) -> tuple[str, str | None]:
    """Pick the separator that splits the file most consistently.

    Federal data isn't always comma-separated (FEC uses "|", BLS uses
    tabs). The stdlib sniffer goes by consistency alone, so a pipe-
    delimited file whose every name reads "LAST, FIRST" looks like a
    perfectly consistent two-column CSV — collapsing sixteen columns into
    two. Among separators that split evenly, the one producing more
    columns is the real one.
    """
    sample = "\n".join(text[:65536].splitlines()[:50])
    if not sample:
        return ",", None

    scores: dict[str, tuple[float, int]] = {}
    for candidate in (",", "|", "\t", ";"):
        try:
            rows = [
                row
                for row in csv.reader(io.StringIO(sample), delimiter=candidate)
                if row
            ]
        except csv.Error:
            continue
        if len(rows) < 2:
            continue
        widths = [len(row) for row in rows]
        common = max(set(widths), key=widths.count)
        if common < 2:
            continue
        scores[candidate] = (widths.count(common) / len(widths), common)

    if not scores:
        return ",", None
    # Most consistent wins; ties go to whichever finds more structure.
    #
    # "comma=2, pipe=3" is genuinely ambiguous — it fits a pipe file whose
    # names read "LAST, FIRST" and a comma file with pipes inside one
    # field. Federal exports are full of the former, so it wins here, and
    # the runner-up is reported so the choice isn't made silently.
    chosen = max(scores, key=lambda d: scores[d])
    rival = next(
        (
            candidate
            for candidate in sorted(scores, key=lambda d: scores[d], reverse=True)
            if candidate != chosen and scores[candidate][0] >= scores[chosen][0]
        ),
        None,
    )
    return chosen, rival


def mixed_delimiter_headings(sheet: Sheet) -> str | None:
    """The separator this sheet might really use, when the split was a
    genuine toss-up.

    "comma=2, pipe=3" fits both a pipe file whose names read "LAST, FIRST"
    and a comma file with pipes inside one field, and nothing in the text
    settles it. Guessing wrong pushes the first record into the heading
    row, where redaction never reaches, so the toss-up is shown rather
    than resolved silently. Reported only when a rival really scored as
    well — a warning that fires on files that parsed perfectly is one
    people learn to click past.
    """
    return sheet.rival_delimiter


def _read_csv(data: bytes) -> Sheet:
    text = _decode(data)
    head = text.lstrip()[:200].lower()
    if head.startswith("<!doctype html") or head.startswith("<html") or "<body" in head:
        raise ValueError(
            "This looks like a saved web page, not a spreadsheet. It usually means "
            "a download failed or a login page was saved by mistake — try "
            "downloading the file again."
        )
    delimiter, rival = _sniff_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if row]
    if rows and has_header(rows[0], rows[1:]):
        headers, body = rows[0], rows[1:]
    else:
        headers = [f"column_{i + 1}" for i in range(len(rows[0]) if rows else 0)]
        body = rows
    body = [[_cell(v) for v in row] for row in body]
    headers, body = _normalize(headers, body)
    return Sheet(
        name="Sheet1",
        headers=headers,
        rows=body,
        delimiter=delimiter,
        rival_delimiter=rival,
    )


def _read_xlsx(data: bytes) -> list[Sheet]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        # read_only defers the real parsing to iter_rows, so pulling the
        # rows here is what actually surfaces a damaged workbook.
        parsed = [
            (ws.title, [[_cell(v) for v in row] for row in ws.iter_rows(values_only=True)])
            for ws in wb.worksheets
        ]
    except Exception:
        # openpyxl raises a wide variety of low-level errors (zip, XML, key)
        # for damaged or password-protected workbooks.
        raise ValueError(
            "This Excel file could not be opened. It may be damaged, "
            "password-protected, or not really an Excel file. Try opening it in "
            "Excel and saving a fresh copy."
        )
    sheets = []
    header_rows: dict[str, int] = {}
    for title, rows in parsed:
        if rows and has_header(rows[0], rows[1:]) and any(c.strip() for c in rows[0]):
            header_rows[title] = 1
            headers, body = rows[0], rows[1:]
        else:
            # A spreadsheet can be headerless too, and its first record
            # would otherwise sit in the header row, which is never redacted.
            # An all-blank first row means uncached formulas, not a
            # heading — treating it as one would lose that record.
            header_rows[title] = 0
            headers = [f"column_{i + 1}" for i in range(len(rows[0]) if rows else 0)]
            body = rows
        headers, body = _normalize(headers, body)
        sheets.append(Sheet(name=title, headers=headers, rows=body))

    if _contains_formulas(data):
        _fill_from_formulas(data, sheets, header_rows)
    return sheets


def _contains_formulas(data: bytes) -> bool:
    """Cheap check of the workbook's XML before paying for a second read.
    Most files have no formulas at all."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    if b"<f" in zf.read(name):
                        return True
    except Exception:
        return False
    return False


def _fill_from_formulas(
    data: bytes, sheets: list[Sheet], header_rows: dict[str, int]
) -> None:
    """A workbook saved by a script or a web exporter often has no cached
    results for its formulas, so those cells read as empty — and a column
    that reads empty later looks deletable. Fall back to the formula text
    for exactly the cells that came back blank.
    """
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    except Exception:
        return
    by_name = {ws.title: ws for ws in wb.worksheets}
    for sheet in sheets:
        ws = by_name.get(sheet.name)
        if ws is None:
            continue
        raw = [[_cell(v) for v in row] for row in ws.iter_rows(values_only=True)]
        # Data starts after the heading row, if this sheet had one.
        offset = header_rows.get(sheet.name, 1)
        for row_number, row in enumerate(sheet.rows, start=offset):
            if row_number >= len(raw):
                break
            source = raw[row_number]
            for i, cell in enumerate(row):
                if not cell.strip() and i < len(source) and source[i].startswith("="):
                    row[i] = source[i]
