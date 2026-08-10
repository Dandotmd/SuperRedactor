"""Serialize sheets back to CSV or XLSX bytes (values only)."""

import csv
import io

from openpyxl import Workbook

from app.engine.readers import Sheet
from app.engine.safety import neutralize

# Excel refuses sheet names longer than this or containing these characters.
MAX_SHEET_NAME = 31
_ILLEGAL_SHEET_CHARS = str.maketrans({c: "-" for c in r"[]:*?/\'"})


def write_csv(sheet: Sheet) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([neutralize(h) for h in sheet.headers])
    writer.writerows([neutralize(c) for c in row] for row in sheet.rows)
    return buf.getvalue().encode("utf-8")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = (name or "Sheet").translate(_ILLEGAL_SHEET_CHARS)[:MAX_SHEET_NAME]
    cleaned = cleaned.strip() or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        tail = f" ({suffix})"
        candidate = cleaned[: MAX_SHEET_NAME - len(tail)] + tail
        suffix += 1
    used.add(candidate)
    return candidate


def write_xlsx(sheets: list[Sheet]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for sheet in sheets:
        ws = wb.create_sheet(title=_safe_sheet_name(sheet.name, used))
        for values in [sheet.headers, *sheet.rows]:
            ws.append(values)
            # openpyxl stores a leading "=" as a live formula. Force text so
            # opening the file can never execute what was in the data.
            for cell in ws[ws.max_row]:
                if cell.data_type == "f":
                    cell.data_type = "s"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
