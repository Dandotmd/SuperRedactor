"""Regressions for inputs that crashed the app or corrupted data.

Each case was found by adversarial testing against real-world file quirks:
Mac line endings, pasted documents in a cell, Excel's "Unicode Text"
export, filenames with smart punctuation, and hostile spreadsheet formulas.
"""

import io

import pytest
from openpyxl import load_workbook

from app.engine.cleaners import analyze, apply_fixes
from app.engine.readers import Sheet, read_file
from app.engine.writers import write_csv, write_xlsx


def by_kind(findings, kind):
    return [f for f in findings if f.kind == kind]


# ---- parsing quirks -------------------------------------------------------

def test_classic_mac_cr_only_line_endings():
    data = b"name,score\rAda,88\rGrace,91\r"
    sheets = read_file("mac.csv", data)
    assert sheets[0].headers == ["name", "score"]
    assert sheets[0].rows == [["Ada", "88"], ["Grace", "91"]]


def test_cell_larger_than_the_csv_field_limit():
    big = "x" * 200_000
    data = f"name,notes\nAda,{big}\n".encode()
    sheets = read_file("big.csv", data)
    assert sheets[0].rows[0][1] == big


def test_utf16_csv_is_decoded_not_mangled():
    data = "name,city\nJosé,Köln\n".encode("utf-16")
    sheets = read_file("x.csv", data)
    assert sheets[0].headers == ["name", "city"]
    assert sheets[0].rows == [["José", "Köln"]]


def test_utf16_without_bom_is_decoded():
    data = "name,city\nJosé,Köln\n".encode("utf-16-le")
    sheets = read_file("x.csv", data)
    assert sheets[0].headers == ["name", "city"]


# ---- formulas must never execute in output --------------------------------

FORMULAS = ["=cmd|'/c calc'!A1", "@SUM(1+1)", "+SUM(A1)", "-cmd|calc"]


def test_csv_output_never_emits_an_executable_cell():
    from app.engine.safety import is_formula_risk

    sheet = Sheet(name="S", headers=["note"], rows=[[f] for f in FORMULAS])
    text = write_csv(sheet).decode()
    for line in text.splitlines()[1:]:
        cell = line.strip('"')
        assert not is_formula_risk(cell), cell
        assert cell.startswith("'"), cell


def test_xlsx_output_stores_formulas_as_text_not_live_formulas():
    sheet = Sheet(name="S", headers=["note"], rows=[[f] for f in FORMULAS])
    wb = load_workbook(io.BytesIO(write_xlsx([sheet])))
    ws = wb["S"]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            assert cell.data_type != "f", f"{cell.value} is a live formula"


def test_negative_numbers_survive_unharmed():
    sheet = Sheet(name="S", headers=["amount"], rows=[["-5"], ["-12.5"]])
    lines = write_csv(sheet).decode().splitlines()
    assert lines[1:] == ["-5", "-12.5"]


ORDINARY_VALUES = [
    "-1.5e10",              # scientific notation
    "+1 (555) 010-0100",    # US phone
    "+44 20 7946 0000",     # international phone
    "-1,234.56",            # negative with thousands separator
    "-$40.00",              # negative currency
    "@example.com",         # email fragment / social handle
    "@danielrupp",
    "-1/2",                 # fraction
    "-",                    # dash placeholder
    "--",
    "-Q3 result",           # a leading dash on prose
    "-Smith (deceased)",    # notes columns are full of these
    "- N/A (not collected)",
    "- see notes (p. 3)",
    "+1 (555) 123-4567 ext. x12",
    "+1 (800) FLOWERS",
    "-Late (excused)",
    "- Urgent!",
    "@johndoe (Twitter)",
]


def test_ordinary_values_are_not_treated_as_formulas():
    from app.engine.safety import is_formula_risk

    for value in ORDINARY_VALUES:
        assert not is_formula_risk(value), value


def test_ordinary_values_are_written_out_unchanged():
    sheet = Sheet(name="S", headers=["v"], rows=[[v] for v in ORDINARY_VALUES])
    written = write_csv(sheet).decode()
    assert "'" not in written, written


def test_command_payloads_are_still_caught():
    from app.engine.safety import is_formula_risk

    for value in [
        "=cmd|'/c calc'!A1",
        "@SUM(1+1)",
        "+SUM(1+1)",
        "-2+3+cmd|'/C calc'!A0",
        "=HYPERLINK(\"http://x\")",
        "@cmd|calc",
        "\t=cmd|calc",
        # Whitespace before the payload must not slip past
        " =cmd|'/c calc'!A1",
        " =cmd|calc",   # non-breaking space
        "​=cmd|calc",   # zero-width space
        "\v=cmd|calc",
        "\f=cmd|calc",
        " =SUM(A1)",    # figure space
    ]:
        assert is_formula_risk(value), repr(value)


def test_negative_currency_is_recognized_as_a_number():
    from app.engine.values import DECORATED_NUMBER, strip_number

    for value, expected in [("-$1,000.50", "-1000.50"), ("$-40.00", "-40.00")]:
        assert DECORATED_NUMBER.match(value), value
        assert strip_number(value) == expected


# ---- cleaners must not delete real data -----------------------------------

def test_sparse_trailing_rows_are_kept():
    # A newly enrolled student with only a name filled in is not a footnote
    sheets = [
        Sheet(
            name="S",
            headers=["Student", "Grade", "Score"],
            rows=[
                ["Ada", "A", "90"],
                ["Grace", "B", "80"],
                ["Alan", "A", "95"],
                ["Kay", "", ""],
            ],
        )
    ]
    cleaned = apply_fixes(sheets, {f.id for f in analyze(sheets)})
    assert [r[0] for r in cleaned[0].rows] == ["Ada", "Grace", "Alan", "Kay"]


def test_keyword_footer_rows_are_still_removed():
    sheets = [
        Sheet(
            name="S",
            headers=["Student", "Score"],
            rows=[["Ada", "90"], ["Grace", "80"], ["Total", "170"]],
        )
    ]
    cleaned = apply_fixes(sheets, {f.id for f in analyze(sheets)})
    assert [r[0] for r in cleaned[0].rows] == ["Ada", "Grace"]


def test_country_code_na_is_not_treated_as_missing():
    sheets = [
        Sheet(
            name="S",
            headers=["Student", "Region"],
            rows=[["Ada", "NA"], ["Grace", "GB"], ["Alan", "NA"]],
        )
    ]
    findings = analyze(sheets)
    assert by_kind(findings, "missing_values") == []
    cleaned = apply_fixes(sheets, {f.id for f in findings})
    assert [r[1] for r in cleaned[0].rows] == ["NA", "GB", "NA"]
    assert cleaned[0].headers == ["Student", "Region"]


def test_none_as_a_category_is_not_treated_as_missing():
    sheets = [
        Sheet(
            name="S",
            headers=["Student", "Allergy"],
            rows=[["Ada", "None"], ["Grace", "Peanut"], ["Alan", "None"]],
        )
    ]
    assert by_kind(analyze(sheets), "missing_values") == []


def test_unambiguous_missing_markers_still_normalized():
    sheets = [
        Sheet(
            name="S",
            headers=["Student", "Score"],
            rows=[["Ada", "N/A"], ["Grace", "NULL"], ["Alan", "88"]],
        )
    ]
    findings = analyze(sheets)
    assert by_kind(findings, "missing_values")[0].count == 2


# ---- excel formulas with no cached value ----------------------------------

def test_partly_formula_column_keeps_every_row():
    # A column where only some cells are formulas lost those cells entirely
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Student", "Score", "Total"])
    ws.append(["Ada", 2, "=B2*10"])
    ws.append(["Bo", 3, 60])
    ws.append(["Cy", 4, "=B4*10"])
    buf = io.BytesIO()
    wb.save(buf)

    sheets = read_file("calc.xlsx", buf.getvalue())
    assert [r[2] for r in sheets[0].rows] == ["=B2*10", "60", "=B4*10"]


def test_xlsx_formula_without_cached_value_reads_as_its_formula():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Student", "Total"])
    ws.append(["Ada", "=1+1"])
    ws.append(["Grace", "=2+2"])
    buf = io.BytesIO()
    wb.save(buf)

    sheets = read_file("calc.xlsx", buf.getvalue())
    assert [r[1] for r in sheets[0].rows] == ["=1+1", "=2+2"]


# ---- performance ----------------------------------------------------------

def test_analyze_is_fast_on_a_large_file():
    import time

    rows = [
        [f"Person {i}", "3/12/2024", f"{i}", "note", "x"] for i in range(50_000)
    ]
    sheets = [
        Sheet(name="S", headers=["name", "seen", "n", "note", "z"], rows=rows)
    ]
    start = time.monotonic()
    analyze(sheets)
    assert time.monotonic() - start < 8.0
