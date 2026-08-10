"""Regressions from the fourth review round."""

import io
import json

from openpyxl import Workbook

from app.engine.detect import suggest_type
from app.engine.leakcheck import find_leaks
from app.engine.readers import Sheet, read_file
from app.engine.redactor import redact
from app.engine.standardize import make_template


# ---- the check must describe the file you actually download ---------------

def test_the_warning_shown_matches_the_file_downloaded():
    """The check runs the redaction to see what happens. Without a shared
    seed it would be measuring a different random run than the download."""
    import io
    import zipfile

    from faker import Faker
    from fastapi.testclient import TestClient

    from app.main import app

    client = TestClient(app)
    faker = Faker()
    pool: set[str] = set()
    while len(pool) < 400:
        pool.add(faker.first_name())
    names = sorted(pool)

    csv = "Student First,Guardian First\n" + "\n".join(
        f"{names[i]},{names[i + 200]}" for i in range(200)
    )
    session = client.post(
        "/api/upload",
        files={"file": ("roster.csv", io.BytesIO(csv.encode()), "text/csv")},
    ).json()["session_id"]
    config = {
        "Sheet1": {"Student First": "first_name", "Guardian First": "first_name"}
    }

    checked = client.post(
        "/api/redact/check", json={"session_id": session, "config": config}
    ).json()
    resp = client.post("/api/redact", json={"session_id": session, "config": config})
    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    out = zf.read([n for n in zf.namelist() if n.endswith(".csv")][0]).decode()
    produced = {c.strip() for line in out.splitlines()[1:] for c in line.split(",")}

    if produced & set(names):
        assert checked["weak_columns"], (
            f"{len(produced & set(names))} real names in the download, "
            f"with a clean check beforehand"
        )


# ---- a removed value must not come back, whatever the case ----------------

def test_uppercase_codes_from_a_removed_column_do_not_come_back():
    """Identifier columns compare case-sensitively, so a removed value has
    to be protected under that comparison too — not only casefolded."""
    # Both columns hold the same shape — one uppercase letter and a digit —
    # so every replacement is drawn from the same 260 possibilities the
    # removed column's values live in.
    kept = [f"{letter}{digit}" for letter in "KLMNOPQRST" for digit in "0123456789"]
    removed = [f"{letter}{digit}" for letter in "ABCDEFGHIJ" for digit in "0123456789"]
    rows = [[kept[i], removed[i]] for i in range(100)]
    sheets = [Sheet(name="S", headers=["Room Code", "Old Room Code"], rows=rows)]
    config = {"S": {"Room Code": "format_preserving", "Old Room Code": "drop"}}

    from app.engine.leakcheck import find_weak_columns

    redacted, _ = redact(sheets, config, seed=5)
    gone = set(removed)
    produced = {c.strip() for row in redacted[0].rows for c in row if c.strip()}
    if produced & gone:
        # 100 values needing replacements out of 260 possible, with 200
        # already spoken for, genuinely cannot be satisfied — but it must
        # never happen quietly.
        assert find_weak_columns(sheets, config, seed=5), (
            f"{len(produced & gone)} removed codes came back with no warning"
        )


def test_uppercase_codes_removed_alongside_a_number_column():
    values = [str(n) for n in range(100, 400)]
    rows = [[str(1000 + i), values[i]] for i in range(300)]
    sheets = [Sheet(name="S", headers=["Count", "Old Count"], rows=rows)]
    config = {"S": {"Count": "number", "Old Count": "drop"}}

    redacted, _ = redact(sheets, config, seed=6)
    removed = set(values)
    produced = {c.strip() for row in redacted[0].rows for c in row if c.strip()}
    assert not (produced & removed), sorted(produced & removed)[:6]


# ---- N1: delimiter sniffing -----------------------------------------------

def test_pipe_delimited_file_whose_names_all_contain_a_comma():
    # "LAST, FIRST" gives every row the same comma count, which fooled the
    # sniffer into reading a 16-column file as 2 columns
    rows = [f"H0AK0010{i}|LAMB, THOMAS|NNE|2020|AK" for i in range(9)]
    data = ("\n".join(rows) + "\n").encode()
    sheets = read_file("cn.csv", data)
    assert len(sheets[0].headers) == 5
    assert len(sheets[0].rows) == 9
    assert sheets[0].rows[0][1] == "LAMB, THOMAS"


def test_an_ambiguous_split_is_flagged_rather_than_trusted():
    """'comma=2, pipe=3' fits two different files and nothing in the text
    settles it. When the headings still hold the other separator, the file
    is flagged so the wrong guess can't pass unnoticed."""
    from app.engine.readers import mixed_delimiter_headings

    for data in (
        b"Ada Lovelace,red|green|blue\nGrace Hopper,a|b|c\nAlan Turing,x|y|z\n",
        b"id,tags|more|cols\n1,red|green|blue\n2,a|b|c\n",
    ):
        sheets = read_file("tags.csv", data)
        assert mixed_delimiter_headings(sheets[0]) == ",", data


def test_an_unambiguous_split_is_not_flagged():
    from app.engine.readers import mixed_delimiter_headings

    for data in (
        b"name,email\nAda,ada@x.org\nBo,bo@x.org\nCy,cy@x.org\n",
        b"name\temail\nAda\tada@x.org\nBo\tbo@x.org\n",
        b"a,b,c\n1,2,3\n4,5,6\n",
    ):
        sheets = read_file("x.csv", data)
        assert mixed_delimiter_headings(sheets[0]) is None, data


def test_the_flag_survives_being_passed_to_the_next_tool():
    """The notice is the mitigation for a wrong split, so it must not
    disappear on the handoff the app encourages."""
    from app.engine.cleaners import clean
    from app.engine.readers import mixed_delimiter_headings
    from app.engine.redactor import redact

    data = b"Ada Lovelace,red|green|blue\nGrace Hopper,a|b|c\nAlan Turing,x|y|z\n"
    sheets = read_file("tags.csv", data)
    assert mixed_delimiter_headings(sheets[0]) == ","

    cleaned, _ = clean(sheets, enabled=set())
    assert mixed_delimiter_headings(cleaned[0]) == ","

    redacted, _ = redact(sheets, {})
    assert mixed_delimiter_headings(redacted[0]) == ","


def test_narrow_pipe_file_with_commas_in_names():
    # Three columns, one comma per row from "LAST, FIRST" — the shape that
    # a "must be twice as wide" rule handed to the comma parser
    data = (
        b"H0AK00105|LAMB, THOMAS|AK\n"
        b"H0AK00113|YOUNG, DONALD E|AK\n"
        b"H0AL01055|CARL, JERRY LEE JR|AL\n"
        b"H0AL01063|CASTORANI, JOHN|AL\n"
    )
    sheets = read_file("roster.txt", data)
    assert len(sheets[0].headers) == 3, sheets[0].headers
    assert len(sheets[0].rows) == 4
    assert sheets[0].rows[0][1] == "LAMB, THOMAS"


def test_five_column_pipe_file_with_two_commas_per_row():
    data = (
        b"1001|SMITH, JOHN|111-22-3333|Boston, MA|02101\n"
        b"1002|DOE, JANE|222-33-4444|Austin, TX|73301\n"
        b"1003|ROE, RICH|333-44-5555|Denver, CO|80014\n"
    )
    sheets = read_file("roster.csv", data)
    assert len(sheets[0].headers) == 5, sheets[0].headers
    assert len(sheets[0].rows) == 3


def test_semicolon_file_with_commas_in_fields():
    data = (
        b"1001;SMITH, JOHN;Boston\n"
        b"1002;DOE, JANE;Austin\n"
        b"1003;ROE, RICH;Denver\n"
    )
    sheets = read_file("x.csv", data)
    assert len(sheets[0].headers) == 3, sheets[0].headers


def test_plain_comma_file_is_still_comma_delimited():
    data = b"name,email\nAda,ada@x.org\nGrace,grace@x.org\n"
    sheets = read_file("x.csv", data)
    assert sheets[0].headers == ["name", "email"]


# ---- headerless files whose first column is numeric ------------------------

HEADERLESS_NUMERIC = [
    b"1001,Ada Lovelace,Ms. Smith,Grade 5\n"
    b"1002,Alan Turing,Mr. Jones,Grade 4\n"
    b"1003,Grace Hopper,Ms. Smith,Grade 6\n",
    b"$1234,Ada Lovelace,Diabetes,Approved\n"
    b"$2345,Alan Turing,Asthma,Approved\n"
    b"$3456,Grace Hopper,Diabetes,Denied\n",
    b"12.50,Ada Lovelace,Diabetes,Approved\n"
    b"13.75,Alan Turing,Asthma,Approved\n"
    b"14.20,Grace Hopper,Diabetes,Denied\n",
    b"2001,Ada Lovelace,Ms. Smith,Grade 5\n"
    b"2002,Alan Turing,Mr. Jones,Grade 4\n"
    b"2003,Grace Hopper,Ms. Smith,Grade 6\n",
]


def test_a_numeric_first_column_does_not_make_record_one_a_heading():
    for data in HEADERLESS_NUMERIC:
        sheets = read_file("roster.csv", data)
        assert sheets[0].headers[0].startswith("column_"), sheets[0].headers
        assert len(sheets[0].rows) == 3, data
        assert "Ada Lovelace" not in sheets[0].headers


def test_the_same_shape_in_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for i, name in enumerate(["Ada Lovelace", "Alan Turing", "Grace Hopper"]):
        ws.append([1001 + i, name, "Ms. Smith", "Grade 5"])
    buf = io.BytesIO()
    wb.save(buf)

    sheets = read_file("roster.xlsx", buf.getvalue())
    assert sheets[0].headers[0].startswith("column_")
    assert len(sheets[0].rows) == 3


def test_year_named_columns_are_still_headings():
    data = b"Name,2023,2024\nAda,88,91\nBo,75,80\nCy,95,99\n"
    sheets = read_file("years.csv", data)
    assert sheets[0].headers == ["Name", "2023", "2024"]


def test_ordinary_headings_over_numeric_columns_are_still_headings():
    data = b"Name,Score,Rank\nAda,88,1\nBo,75,2\nCy,95,3\n"
    sheets = read_file("x.csv", data)
    assert sheets[0].headers == ["Name", "Score", "Rank"]


# ---- N2: headerless xlsx --------------------------------------------------

def test_headerless_xlsx_is_detected():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for i in range(4):
        ws.append([f"H0AK0010{i}", "LAMB THOMAS", "AK", "111-22-3333"])
    buf = io.BytesIO()
    wb.save(buf)

    sheets = read_file("cn.xlsx", buf.getvalue())
    assert sheets[0].headers[0] == "column_1"
    assert len(sheets[0].rows) == 4


def test_formula_backfill_lines_up_on_a_headerless_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for i in range(1, 4):
        ws.append([f"H0AK0010{i}", f"NAME{i}", f"=A{i}&B{i}"])
    buf = io.BytesIO()
    wb.save(buf)

    sheets = read_file("cn.xlsx", buf.getvalue())
    assert [r[2] for r in sheets[0].rows] == ["=A1&B1", "=A2&B2", "=A3&B3"]


def test_xlsx_with_a_real_header_row_keeps_it():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Name", "Email", "Score"])
    ws.append(["Ada", "ada@x.org", 88])
    ws.append(["Grace", "grace@x.org", 91])
    buf = io.BytesIO()
    wb.save(buf)

    sheets = read_file("x.xlsx", buf.getvalue())
    assert sheets[0].headers == ["Name", "Email", "Score"]


# ---- N3: values ending in punctuation --------------------------------------

def test_a_name_ending_in_a_full_stop_is_found_in_a_sentence():
    sheets = [
        Sheet(
            name="S",
            headers=["Client", "Notes"],
            rows=[
                ["Ryan Hall Jr.", "spoke to Ryan Hall Jr. today"],
                ["Ada Lovelace", "spoke to Ada Lovelace today"],
                ["Bo Diddley", "nothing"],
            ],
        )
    ]
    leaks = find_leaks(sheets, {"S": {"Client": "person_name"}})
    assert leaks and leaks[0].count == 2, leaks


# ---- N4: removed columns as leak sources -----------------------------------

def test_a_value_from_a_removed_column_quoted_elsewhere_is_reported():
    sheets = [
        Sheet(
            name="S",
            headers=["Student", "Notes"],
            rows=[
                ["Ida Wells", "Ida Wells called"],
                ["Bo Diddley", "nothing"],
                ["Cy Young", "nothing"],
            ],
        )
    ]
    assert find_leaks(sheets, {"S": {"Student": "drop"}})


# ---- N5: case is data in identifier columns --------------------------------

def test_case_distinct_identifiers_stay_distinct():
    sheets = [
        Sheet(name="S", headers=["code"], rows=[["aB3xK9"], ["Ab3Xk9"], ["ZZ9zz1"]])
    ]
    out, mapping = redact(sheets, {"S": {"code": "format_preserving"}})
    produced = [r[0] for r in out[0].rows]
    assert produced[0] != produced[1], "two different IDs must not share a fake"
    assert len(mapping["S"]["code"]) == 3


def test_case_variants_of_a_name_still_share_a_fake():
    sheets = [
        Sheet(name="S", headers=["name"], rows=[["Ada Lovelace"], ["ADA LOVELACE"]])
    ]
    out, _ = redact(sheets, {"S": {"name": "person_name"}})
    assert out[0].rows[0][0] == out[0].rows[1][0]


# ---- N7: role words must not over-trigger ----------------------------------

def test_role_word_with_a_non_name_qualifier_is_not_a_name():
    for header in (
        "Employee Number", "Patient Number", "Client Code", "Staff Count",
        "Customer Segment", "Teacher Rating", "Student Count",
    ):
        assert suggest_type(header) != "person_name", header


def test_plain_role_words_are_still_names():
    for header in ("Patient", "Student", "Guardian", "Emergency Contact"):
        assert suggest_type(header) == "person_name", header


# ---- templates: remembering values is opt-in -------------------------------

def test_the_template_the_api_returns_carries_no_real_values():
    """The saved file is whatever the API calls the template. Offering
    candidate values in the same object put them in the shared file."""
    import io

    from fastapi.testclient import TestClient

    from app.main import app

    client = TestClient(app)
    csv = (
        b"Status,Diagnosis\n"
        b"Active,Asthma\nInactive,Diabetes\nActive,Asthma\n"
        b"Inactive,Diabetes\nActive,Asthma\nInactive,Diabetes\n"
    )
    session = client.post(
        "/api/upload", files={"file": ("roster.csv", io.BytesIO(csv), "text/csv")}
    ).json()["session_id"]
    body = client.post(
        "/api/standardize/template", json={"session_id": session}
    ).json()

    saved = json.dumps(body["template"])
    for value in ("Asthma", "Diabetes", "Active", "Inactive"):
        assert value not in saved, f"{value} reached the template file"
    # the candidates are still offered, just not inside the template
    assert body["suggested_values"]["Diagnosis"] == ["Asthma", "Diabetes"]


def test_templates_do_not_remember_values_by_default():
    sheet = Sheet(
        name="S",
        headers=["Program", "Dx", "Rx", "ICD10"],
        rows=[
            ["Housing", "Asthma", "Albuterol", "J45"],
            ["Care", "Diabetes", "Metformin", "E11"],
            ["Housing", "Asthma", "Albuterol", "J45"],
            ["Care", "Diabetes", "Metformin", "E11"],
            ["Housing", "Asthma", "Albuterol", "J45"],
            ["Care", "Diabetes", "Metformin", "E11"],
        ],
    )
    template = make_template(sheet, name="t")
    for column in template["columns"]:
        assert "values" not in column, column["name"]
    # the columns that could remember a list are offered, not applied
    assert set(template["can_remember_values"]) >= {"Program", "Dx", "Rx", "ICD10"}
